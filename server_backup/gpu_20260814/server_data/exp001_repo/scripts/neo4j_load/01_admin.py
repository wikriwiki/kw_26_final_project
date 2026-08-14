"""District·Dong 노드 + HAS_DONG + ADJACENT_TO 적재.

입력:
  data/neo4j_load/admin/KIKcd_H.xlsx   # 행정동코드, 시도명, 시군구명, 읍면동명, 생성일자, 말소일자
  data/neo4j_load/pois/소상공인...서울_202603.csv  # commerce CSV에서 동별 좌표 평균 산출

출력:
  Neo4j:
    (:District {code, name})
    (:Dong {code, name, lon, lat})
    (:District)-[:HAS_DONG]->(:Dong)
    (:Dong)-[:ADJACENT_TO]-(:Dong)   # 좌표 거리 ≤ 1.5km

서울만 적재 (코드 prefix "11").
"""
from __future__ import annotations

import csv
import math
import sys
import warnings
from collections import defaultdict
from pathlib import Path

warnings.filterwarnings("ignore")
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _common import driver_session, bulk_run

PROJECT_ROOT = Path(__file__).resolve().parents[2]
ADMIN_XLSX = PROJECT_ROOT / "data" / "neo4j_load" / "admin" / "KIKcd_H.xlsx"
COMMERCE_CSV = PROJECT_ROOT / "data" / "neo4j_load" / "pois" / "소상공인시장진흥공단_상가(상권)정보_서울_202603.csv"

# commerce CSV 컬럼 인덱스 (이전 검증)
COL_C_DONG_CD = 15  # 행정동코드 (NSO 8자리)
COL_C_LON = 37
COL_C_LAT = 38

ADJACENCY_KM = 1.5  # 인접 동 판정 거리


def parse_kik_admin() -> tuple[list[dict], list[dict], list[tuple[str, str]]]:
    """KIKcd_H.xlsx → (districts, dongs, has_dong_pairs).

    행정동코드 10자리 MOPAS 구조:
        시도(2) + 시군구(3) + 동(3) + 통(2)
        - 1100000000 = 서울 시도
        - 1111000000 = 종로구 (시군구)
        - 1111053000 = 사직동 (행정동, 마지막 5자리 != 0)

    Dong code는 NSO 8자리 (= MOPAS[:8]).
    District code는 자치구 5자리 코드 (= MOPAS[:5]).
    """
    wb = load_workbook(ADMIN_XLSX, read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]

    districts: dict[str, str] = {}    # code5 -> name
    dongs: dict[str, dict] = {}       # code8 -> {code, name, district_code5}
    for row in ws.iter_rows(values_only=True, min_row=2):
        if not row or row[0] is None:
            continue
        code10 = str(row[0])
        sido = row[1]
        sigungu = row[2]
        dong_name = row[3]
        malsoji = row[5]  # 말소일자
        if malsoji:
            continue  # 폐지된 행정동 제외
        # 서울만
        if not code10.startswith("11"):
            continue
        # 자치구 행: 1111000000 (sigungu only)
        if sigungu and not dong_name:
            districts[code10[:5]] = sigungu
            continue
        # 행정동 행: 1111053000
        if dong_name:
            code8 = code10[:8]
            district5 = code10[:5]
            dongs[code8] = {
                "code": code8,
                "name": dong_name,
                "sigungu_name": sigungu or "",
                "district_code5": district5,
            }

    district_list = [{"code": c, "name": n} for c, n in sorted(districts.items())]
    dong_list = list(dongs.values())
    has_dong = [(d["district_code5"], d["code"]) for d in dong_list]
    return district_list, dong_list, has_dong


def compute_dong_centroids(dong_codes_set: set[str]) -> dict[str, tuple[float, float]]:
    """commerce CSV에서 동별 (lon, lat) 평균 → 중심좌표."""
    sums = defaultdict(lambda: [0.0, 0.0, 0])  # code -> [sum_lon, sum_lat, count]
    with open(COMMERCE_CSV, encoding="utf-8-sig") as f:
        r = csv.reader(f)
        next(r, None)
        for row in r:
            if len(row) <= COL_C_LAT:
                continue
            code = row[COL_C_DONG_CD]
            if code not in dong_codes_set:
                continue
            try:
                lon = float(row[COL_C_LON])
                lat = float(row[COL_C_LAT])
            except (ValueError, IndexError):
                continue
            sums[code][0] += lon
            sums[code][1] += lat
            sums[code][2] += 1
    centroids = {}
    for code, (slon, slat, n) in sums.items():
        if n > 0:
            centroids[code] = (slon / n, slat / n)
    return centroids


def haversine_km(lon1, lat1, lon2, lat2):
    R = 6371.0
    p1 = math.radians(lat1)
    p2 = math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 2 * R * math.asin(math.sqrt(a))


def compute_adjacency(dongs_with_coord: list[dict], threshold_km: float) -> list[tuple[str, str]]:
    """모든 동 쌍 O(N^2) 거리 검사. 424^2 ≈ 18만 쌍, ~1초."""
    pairs = []
    n = len(dongs_with_coord)
    for i in range(n):
        a = dongs_with_coord[i]
        for j in range(i + 1, n):
            b = dongs_with_coord[j]
            if haversine_km(a["lon"], a["lat"], b["lon"], b["lat"]) <= threshold_km:
                pairs.append((a["code"], b["code"]))
    return pairs


def main():
    print(f"[parse] reading {ADMIN_XLSX.name}")
    districts, dongs, has_dong = parse_kik_admin()
    print(f"  Districts (Seoul): {len(districts)}")
    print(f"  Dongs (Seoul):     {len(dongs)}")
    print(f"  HAS_DONG pairs:    {len(has_dong)}")

    print(f"[centroid] computing from commerce CSV ...")
    dong_codes_set = {d["code"] for d in dongs}
    centroids = compute_dong_centroids(dong_codes_set)
    print(f"  Dongs with centroid: {len(centroids)} / {len(dongs)}")

    # attach lon/lat to dongs
    for d in dongs:
        if d["code"] in centroids:
            d["lon"], d["lat"] = centroids[d["code"]]
        else:
            d["lon"], d["lat"] = None, None
    dongs_with_coord = [d for d in dongs if d["lon"] is not None]

    print(f"[adjacency] computing ADJACENT_TO (≤ {ADJACENCY_KM}km) ...")
    adj_pairs = compute_adjacency(dongs_with_coord, ADJACENCY_KM)
    print(f"  ADJACENT_TO pairs: {len(adj_pairs)}")

    print(f"[neo4j] writing ...")
    with driver_session() as s:
        # District
        s.run("""
            UNWIND $batch AS d
            MERGE (n:District {code: d.code})
            SET n.name = d.name
        """, batch=districts)
        print(f"  + District x {len(districts)}")

        # Dong
        s.run("""
            UNWIND $batch AS d
            MERGE (n:Dong {code: d.code})
            SET n.name = d.name,
                n.sigungu_name = d.sigungu_name,
                n.lon = d.lon, n.lat = d.lat
        """, batch=dongs)
        print(f"  + Dong x {len(dongs)}")

        # HAS_DONG
        s.run("""
            UNWIND $batch AS p
            MATCH (dis:District {code: p[0]}), (d:Dong {code: p[1]})
            MERGE (dis)-[:HAS_DONG]->(d)
        """, batch=[[a, b] for a, b in has_dong])
        print(f"  + HAS_DONG x {len(has_dong)}")

        # ADJACENT_TO (양방향)
        bulk_run(s, """
            UNWIND $batch AS p
            MATCH (a:Dong {code: p[0]}), (b:Dong {code: p[1]})
            MERGE (a)-[:ADJACENT_TO]->(b)
            MERGE (b)-[:ADJACENT_TO]->(a)
        """, batch=[[a, b] for a, b in adj_pairs])
        print(f"  + ADJACENT_TO x {len(adj_pairs)*2}")

    print()
    print("[done] admin loaded.")


if __name__ == "__main__":
    main()
