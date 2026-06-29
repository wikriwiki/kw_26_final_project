"""K-apt 단지 기본정보 (xlsx) → residence POI CSV.

흐름:
1. xlsx 로드, 서울특별시만 필터 (~3,164건)
2. 도로명주소 우선, 없으면 법정동주소 사용
3. V-WORLD Geocoder로 좌표 + 행정동코드(level4AC) 획득
4. CSV로 저장: id, name, lon, lat, dong_code, dong_name, sigungu, source_addr, geocode_status

출력:
  data/neo4j_load/pois/residence.csv
  data/neo4j_load/pois/_residence_failures.csv (실패 case)
  data/neo4j_load/pois/_residence_summary.json (통계)
"""
from __future__ import annotations

import csv
import json
import sys
import time
import warnings
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

warnings.filterwarnings("ignore")
from openpyxl import load_workbook

# import sibling package
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from geocode.vworld_client import VWorldGeocoder, load_api_key

# === 경로 ===
PROJECT_ROOT = Path(__file__).resolve().parents[2]
XLSX_PATH = PROJECT_ROOT / "data" / "neo4j_load" / "pois" / "20260508_단지_기본정보.xlsx"
OUT_DIR = PROJECT_ROOT / "data" / "neo4j_load" / "pois"
OUT_CSV = OUT_DIR / "residence.csv"
OUT_FAIL = OUT_DIR / "_residence_failures.csv"
OUT_STATS = OUT_DIR / "_residence_summary.json"
CACHE_PATH = PROJECT_ROOT / "scripts" / "geocode" / "cache.sqlite"

# === xlsx 컬럼 인덱스 (검증 완료) ===
COL_SIDO = 0
COL_SIGUNGU = 1
COL_DONG_TEXT = 2   # 읍면
COL_CODE = 4         # 단지코드
COL_NAME = 5         # 단지명
COL_ADDR_JIBUN = 7   # 법정동주소
COL_ADDR_ROAD = 9    # 도로명주소


def load_seoul_records() -> list[dict]:
    """xlsx에서 서울 단지만 추출. header는 row 2 (1-based), 데이터는 row 3+."""
    wb = load_workbook(XLSX_PATH, read_only=False, data_only=True)
    ws = wb["sheet1"]
    records: list[dict] = []
    # skip notice (row 1) + header (row 2)
    for row in ws.iter_rows(values_only=True, min_row=3):
        sido = row[COL_SIDO]
        if sido != "서울특별시":
            continue
        records.append({
            "code": str(row[COL_CODE]) if row[COL_CODE] else "",
            "name": str(row[COL_NAME]) if row[COL_NAME] else "",
            "sigungu": str(row[COL_SIGUNGU]) if row[COL_SIGUNGU] else "",
            "dong_text": str(row[COL_DONG_TEXT]) if row[COL_DONG_TEXT] else "",
            "addr_road": str(row[COL_ADDR_ROAD]).strip() if row[COL_ADDR_ROAD] else "",
            "addr_jibun": str(row[COL_ADDR_JIBUN]).strip() if row[COL_ADDR_JIBUN] else "",
        })
    return records


def _split_first(addr: str) -> str:
    """K-apt 데이터엔 동별 주소가 콤마로 연결된 경우가 있음. 첫 주소만 사용."""
    if not addr:
        return ""
    return addr.split(",")[0].strip()


def choose_address(rec: dict) -> tuple[str, str]:
    """(address, type_first) 결정. 도로명 우선, 없으면 지번. 콤마 다중 주소는 첫 개만."""
    if rec["addr_road"]:
        return _split_first(rec["addr_road"]), "road"
    if rec["addr_jibun"]:
        first = _split_first(rec["addr_jibun"])
        # 지번주소 끝의 "단지명" 같은 텍스트 제거: "서울특별시 성동구 응봉동 100 응봉대림1차" → 마지막 토큰이 한글만이면 제거
        toks = first.split()
        while toks and not any(c.isdigit() or c == "-" for c in toks[-1]):
            toks.pop()
        cleaned = " ".join(toks) if toks else first
        return cleaned.replace("번지", "").strip(), "parcel"
    # fallback: 시군구 + 동 텍스트만으로
    fb = f"서울특별시 {rec['sigungu']} {rec['dong_text']}".strip()
    return fb, "parcel"


def main():
    key = load_api_key()
    print(f"[init] V-WORLD key loaded: {key[:8]}...")
    geo = VWorldGeocoder(key, CACHE_PATH, max_retries=3)

    print(f"[load] reading {XLSX_PATH.name} ...")
    records = load_seoul_records()
    print(f"[load] Seoul records: {len(records)}")

    # Geocoding (concurrent)
    print(f"[geocode] starting batch (concurrency=10) ...")
    t0 = time.time()
    results = [None] * len(records)

    def worker(i_rec):
        i, rec = i_rec
        addr, type_first = choose_address(rec)
        r = geo.geocode_one(addr, type_first=type_first, fallback=True)
        return i, addr, type_first, r

    with ThreadPoolExecutor(max_workers=10) as ex:
        futures = [ex.submit(worker, (i, r)) for i, r in enumerate(records)]
        done = 0
        for fut in as_completed(futures):
            i, addr, type_first, gr = fut.result()
            results[i] = (addr, type_first, gr)
            done += 1
            if done % 200 == 0 or done == len(records):
                elapsed = time.time() - t0
                rate = done / elapsed if elapsed > 0 else 0
                print(f"  progress {done}/{len(records)}  ({rate:.1f} req/s, cache hit {geo.stats['hit']})")

    # write output
    out_rows = []
    fail_rows = []
    sigungu_counter = {}
    status_counter = {"OK": 0, "NOT_FOUND": 0, "ERROR": 0}

    for rec, (addr, type_first, gr) in zip(records, results):
        status_counter[gr.status] = status_counter.get(gr.status, 0) + 1
        if gr.status == "OK" and gr.lon is not None and gr.lat is not None:
            dong_code_nso = (gr.level4AC[:8] if gr.level4AC and len(gr.level4AC) >= 8 else "")
            sigungu_counter[gr.level2 or ""] = sigungu_counter.get(gr.level2 or "", 0) + 1
            out_rows.append({
                "id": f"R_{rec['code']}",
                "name": rec["name"],
                "lon": f"{gr.lon:.7f}",
                "lat": f"{gr.lat:.7f}",
                "dong_code": dong_code_nso,
                "dong_name": gr.level4A or "",
                "sigungu": gr.level2 or "",
                "source_addr": addr,
                "addr_type": gr.type_used,
            })
        else:
            fail_rows.append({
                "id": f"R_{rec['code']}",
                "name": rec["name"],
                "sigungu_in": rec["sigungu"],
                "dong_text_in": rec["dong_text"],
                "addr_road_in": rec["addr_road"],
                "addr_jibun_in": rec["addr_jibun"],
                "tried_addr": addr,
                "status": gr.status,
                "error": gr.raw_error or "",
            })

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["id", "name", "lon", "lat", "dong_code", "dong_name", "sigungu", "source_addr", "addr_type"])
        w.writeheader()
        w.writerows(out_rows)

    with open(OUT_FAIL, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["id", "name", "sigungu_in", "dong_text_in", "addr_road_in", "addr_jibun_in", "tried_addr", "status", "error"])
        w.writeheader()
        w.writerows(fail_rows)

    stats = {
        "input_seoul": len(records),
        "success": len(out_rows),
        "failed": len(fail_rows),
        "status_distribution": status_counter,
        "sigungu_distribution": sigungu_counter,
        "geocode_client_stats": geo.stats,
        "elapsed_sec": round(time.time() - t0, 1),
    }
    OUT_STATS.write_text(json.dumps(stats, ensure_ascii=False, indent=2), encoding="utf-8")

    print()
    print(f"[done] success: {len(out_rows)} / failed: {len(fail_rows)}")
    print(f"[done] sigungu distribution: {dict(sorted(sigungu_counter.items(), key=lambda x: -x[1])[:5])} ...")
    print(f"[done] elapsed: {stats['elapsed_sec']}s")
    print()
    print(f"  -> {OUT_CSV}")
    print(f"  -> {OUT_FAIL}")
    print(f"  -> {OUT_STATS}")


if __name__ == "__main__":
    main()
